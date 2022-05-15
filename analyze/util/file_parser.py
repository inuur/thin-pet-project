from openpyxl import load_workbook


def parse(file, pk):
    workbook = load_workbook(file, keep_vba=True, data_only=True)
    wb_sheet_names = workbook.sheetnames
    first_sheet = workbook[wb_sheet_names[0]]
    analyzes = list()
    b = list()  # ????????
    for row in first_sheet.rows:
        if not b:
            b = [str(i.value).lower() for i in row if i.value]
        else:
            list1 = {b[i]: row[i + 1].value for i in range(len(b))}
            analyzes.append(
                {
                    'name': list1['name'],
                    'k_more0_25': list1['> 0.25'],
                    'k_01_025': list1["0.1-0.25"],
                    'k_005_01': list1["0.05-0.1"],
                    'k_005_001': list1["0.05-0.01"],
                    'k_less_0_01': list1[" < 0.01"],
                    'k_trusk': list1["коэф. по траску"],
                    'med_dim': list1["медианный диаметр"],
                    'st_sort': list1["степень сортированности"],
                    'degree_of_porosity': list1["процент пористости"],
                    'thin_sections_plots_id': pk
                }
            )
    return analyzes
